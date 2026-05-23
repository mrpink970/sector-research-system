#!/usr/bin/env python3
"""
ETF Daily Data Updater - SIMPLIFIED VERSION
Fetches only last 30 days of data
Updated for main repo structure - stores data in data/4_etf/
"""

import sys
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime, timedelta
import numpy as np
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# PATH CONFIGURATION - Updated for main repo structure
# ============================================================================
from pathlib import Path

# Updated path: data/4_etf/4_ETF_Workbook.xlsx
WORKBOOK_PATH = Path("data/4_etf/4_ETF_Workbook.xlsx")

TICKERS = ['SOXL', 'TQQQ', 'SOXS', 'SQQQ']


def fetch_ticker_data(ticker, days=30):
    """Fetch data for a single ticker - FIXED for yfinance version changes"""
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    print(f"  Fetching {ticker} from {start_date.date()} to {end_date.date()}")
    
    try:
        data = yf.download(ticker, start=start_date, end=end_date, progress=False)
        if data.empty:
            print(f"    WARNING: No data for {ticker}")
            return None
        
        # Reset index to get date as column
        data = data.reset_index()
        
        # Handle different column names (yfinance version dependent)
        if 'Date' in data.columns:
            data['Date'] = pd.to_datetime(data['Date']).dt.strftime('%Y-%m-%d')
        elif 'index' in data.columns:
            data['Date'] = pd.to_datetime(data['index']).dt.strftime('%Y-%m-%d')
        else:
            # Use the first column as date (fallback)
            first_col = data.columns[0]
            data['Date'] = pd.to_datetime(data[first_col]).dt.strftime('%Y-%m-%d')
        
        return data
    except Exception as e:
        print(f"    ERROR: {e}")
        return None


def update_workbook(workbook_path):
    """Update the Excel workbook with latest ETF data"""
    print("=" * 60)
    print("ETF DATA UPDATER - SIMPLIFIED")
    print("=" * 60)
    print(f"Workbook path: {workbook_path}")
    
    # Ensure directory exists
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Fetch data for all tickers
    all_data = {}
    for ticker in TICKERS:
        df = fetch_ticker_data(ticker, days=30)
        if df is not None:
            all_data[ticker] = df
    
    if not all_data:
        print("ERROR: No data fetched")
        return False
    
    # Get all unique dates
    all_dates = set()
    for df in all_data.values():
        all_dates.update(df['Date'].tolist())
    all_dates = sorted(all_dates)
    
    print(f"\nDates: {all_dates[0]} to {all_dates[-1]} ({len(all_dates)} days)")
    
    # Build the DataFrame
    rows = []
    for date in all_dates:
        row = {'Date': date}
        for ticker, df in all_data.items():
            ticker_data = df[df['Date'] == date]
            if not ticker_data.empty:
                # Extract scalar values
                open_val = ticker_data['Open'].iloc[0]
                high_val = ticker_data['High'].iloc[0]
                low_val = ticker_data['Low'].iloc[0]
                close_val = ticker_data['Close'].iloc[0]
                
                # Convert to scalar
                if hasattr(open_val, 'item'):
                    open_val = open_val.item()
                    high_val = high_val.item()
                    low_val = low_val.item()
                    close_val = close_val.item()
                
                row[f"{ticker}_Open"] = round(float(open_val), 4) if open_val is not None else None
                row[f"{ticker}_High"] = round(float(high_val), 4) if high_val is not None else None
                row[f"{ticker}_Low"] = round(float(low_val), 4) if low_val is not None else None
                row[f"{ticker}_Close"] = round(float(close_val), 4) if close_val is not None else None
            else:
                row[f"{ticker}_Open"] = None
                row[f"{ticker}_High"] = None
                row[f"{ticker}_Low"] = None
                row[f"{ticker}_Close"] = None
        rows.append(row)
    
    df_final = pd.DataFrame(rows)
    
    # Calculate returns
    for ticker in TICKERS:
        close_col = f"{ticker}_Close"
        if close_col in df_final.columns:
            df_final[close_col] = pd.to_numeric(df_final[close_col], errors='coerce')
            df_final[f"{ticker}_%Chg"] = df_final[close_col].pct_change() * 100
            df_final[f"{ticker}_%Chg"] = df_final[f"{ticker}_%Chg"].round(4)
            df_final[f"{ticker}_3D"] = df_final[close_col].pct_change(periods=3) * 100
            df_final[f"{ticker}_3D"] = df_final[f"{ticker}_3D"].round(4)
            df_final[f"{ticker}_5D"] = df_final[close_col].pct_change(periods=5) * 100
            df_final[f"{ticker}_5D"] = df_final[f"{ticker}_5D"].round(4)
    
    print(f"\nFinal data: {len(df_final)} rows, {len(df_final.columns)} columns")
    
    # Write to Excel - FIXED: Properly handle new/existing workbook
    if workbook_path.exists():
        try:
            wb = load_workbook(workbook_path)
            print("Loaded existing workbook")
        except Exception as e:
            print(f"Could not load existing workbook: {e}")
            print("Creating new workbook")
            wb = Workbook()
    else:
        print("Creating new workbook")
        wb = Workbook()
        # Remove the default sheet that comes with new workbook
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # Remove old Daily_Data sheet if it exists
    if 'Daily_Data' in wb.sheetnames:
        wb.remove(wb['Daily_Data'])
    
    # Create new Daily_Data sheet
    ws = wb.create_sheet('Daily_Data', 0)  # Insert at beginning
    
    # Write headers
    headers = list(df_final.columns)
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data
    rows_written = 0
    for row_idx, row in df_final.iterrows():
        excel_row = row_idx + 2
        for col_idx, header in enumerate(headers, 1):
            value = row[header]
            if pd.notna(value):
                try:
                    ws.cell(row=excel_row, column=col_idx, value=float(value))
                except (TypeError, ValueError):
                    ws.cell(row=excel_row, column=col_idx, value=value)
        rows_written += 1
    
    print(f"Wrote {rows_written} rows to Daily_Data sheet")
    
    # Ensure Signal sheet
    if 'Signal' not in wb.sheetnames:
        ws_signal = wb.create_sheet('Signal')
        ws_signal['D23'] = 'SOXL'
        ws_signal['D24'] = 'TQQQ'
        ws_signal['D27'] = datetime.now().strftime('%Y-%m-%d')
        print("Created Signal sheet")
    else:
        # Update signal date
        ws_signal = wb['Signal']
        ws_signal['D27'] = datetime.now().strftime('%Y-%m-%d')
        print("Updated Signal sheet date")
    
    # Save the workbook
    wb.save(workbook_path)
    print(f"\n✅ Updated {workbook_path}")
    print(f"   Rows: {len(df_final)}")
    
    # Print last few rows for verification
    print("\nLast 3 rows of data:")
    cols_to_show = ['Date', 'SOXL_Open', 'SOXL_High', 'SOXL_Close']
    available_cols = [c for c in cols_to_show if c in df_final.columns]
    if available_cols:
        print(df_final.tail(3)[available_cols].to_string())
    else:
        print("No SOXL data available")
    
    return True


if __name__ == "__main__":
    success = update_workbook(WORKBOOK_PATH)
    sys.exit(0 if success else 1)
